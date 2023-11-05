from services.readerService.fileReader.FileReader import FileReader
import openpyxl

class XlsxReader(FileReader):
    def __init__(self, file):
        self.file = file
    
    def read(self):
        workbook = openpyxl.load_workbook(self.file, data_only=True)
        worksheet = workbook.active
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append({worksheet.cell(
                row=1, column=i).value: value for i, value in enumerate(row, start=1)})

        return data